from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import os
import sqlite3
import pandas as pd
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# Initialize Flask app
app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "static/uploads/"
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Initialize Database (Removed Critical Column)
def init_db():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    # Create table if it doesn't exist
    c.execute('''CREATE TABLE IF NOT EXISTS problems (
                 id INTEGER PRIMARY KEY, category TEXT, description TEXT, 
                 image TEXT, date TEXT, comment TEXT, progress TEXT, priority TEXT DEFAULT 'Medium')''')

    # Check if 'priority' column exists, if not, add it
    c.execute("PRAGMA table_info(problems)")
    columns = [col[1] for col in c.fetchall()]

    if "priority" not in columns:
        c.execute("ALTER TABLE problems ADD COLUMN priority TEXT DEFAULT 'Medium'")

    conn.commit()
    conn.close()

init_db()

# Home Page (Form + Problem List)
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        category = request.form["category"]
        description = request.form["description"]
        comment = request.form["comment"]
        progress = request.form["progress"]
        priority = request.form["priority"]
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Handle Image Upload
        image_filename = None
        if "image" in request.files:
            image = request.files["image"]
            if image.filename:
                image_filename = image.filename
                image.save(os.path.join(app.config["UPLOAD_FOLDER"], image_filename))

        # Save to Database
        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("INSERT INTO problems (category, description, image, date, comment, progress, priority) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  (category, description, image_filename, date, comment, progress, priority))
        conn.commit()
        conn.close()

        return redirect(url_for("index"))

    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT * FROM problems ORDER BY id DESC")
    problems = c.fetchall()
    conn.close()
    return render_template("index.html", problems=problems)

# Edit Page
@app.route("/edit/<int:id>", methods=["GET"])
def edit(id):
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT * FROM problems WHERE id = ?", (id,))
    problem = c.fetchone()
    conn.close()

    if not problem:
        return "Problem not found", 404

    return render_template("edit.html", problem=problem)

# Update Record
@app.route("/update/<int:id>", methods=["POST"])
def update(id):
    category = request.form["category"]
    description = request.form["description"]
    comment = request.form["comment"]
    progress = request.form["progress"]
    priority = request.form["priority"]

    # Handle Image Upload
    image_filename = None
    if "image" in request.files:
        image = request.files["image"]
        if image.filename:
            image_filename = image.filename
            image.save(os.path.join(app.config["UPLOAD_FOLDER"], image_filename))

    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    if image_filename:
        c.execute("UPDATE problems SET category = ?, description = ?, image = ?, comment = ?, progress = ?, priority = ? WHERE id = ?",
                  (category, description, image_filename, comment, progress, priority, id))
    else:
        c.execute("UPDATE problems SET category = ?, description = ?, comment = ?, progress = ?, priority = ? WHERE id = ?",
                  (category, description, comment, progress, priority, id))

    conn.commit()
    conn.close()

    return redirect(url_for("index"))

# Delete Record
@app.route("/delete/<int:id>")
def delete(id):
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("DELETE FROM problems WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    return redirect(url_for("index"))

# Update Progress via AJAX
@app.route("/update_progress/<int:id>", methods=["POST"])
def update_progress(id):
    new_progress = request.form["progress"]

    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("UPDATE problems SET progress = ? WHERE id = ?", (new_progress, id))
    conn.commit()
    conn.close()

    return jsonify({"message": "Progress updated successfully!", "progress": new_progress})

# Export Data to Excel
def export_problems():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()
    c.execute("SELECT * FROM problems")
    problems = c.fetchall()
    conn.close()

    # Create an Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Problem Report"

    # Define headers
    headers = ["ID", "Category", "Description", "Image", "Date", "Comment", "Progress", "Priority"]
    ws.append(headers)

    # Set column widths for better visibility
    column_widths = [5, 15, 30, 20, 20, 30, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Set row height for images
    row_height = 50  # Adjust this value as needed

    # Insert data into the worksheet
    for row_index, problem in enumerate(problems, start=2):
        ws.row_dimensions[row_index].height = row_height  # Set row height

        # Add text data
        ws.cell(row=row_index, column=1, value=problem[0])  # ID
        ws.cell(row=row_index, column=2, value=problem[1])  # Category
        ws.cell(row=row_index, column=3, value=problem[2])  # Description
        ws.cell(row=row_index, column=5, value=problem[4])  # Date
        ws.cell(row=row_index, column=6, value=problem[5] if problem[5] else "No Comment")  # Comment
        ws.cell(row=row_index, column=7, value=problem[6])  # Progress
        ws.cell(row=row_index, column=8, value=problem[7])  # Priority

        # Insert Image
        image_filename = problem[3]  # Image path stored in DB
        if image_filename:
            image_path = os.path.join(app.config["UPLOAD_FOLDER"], image_filename)
            if os.path.exists(image_path):
                img = ExcelImage(image_path)
                img.width, img.height = 100, 75  # Resize image (fixed size)
                ws.add_image(img, f"D{row_index}")  # Place image in column D

    # Save the file
    filepath = "problem_report.xlsx"
    wb.save(filepath)
    return filepath

@app.route("/export")
def export():
    filepath = export_problems()
    return send_file(filepath, as_attachment=True)

# Scheduled Weekly Export (Runs Every Sunday at 11 PM)
scheduler = BackgroundScheduler()
scheduler.add_job(export_problems, 'cron', day_of_week='sun', hour=23, minute=0)
scheduler.start()

if __name__ == "__main__":
    app.run(debug=True)
